from io import BytesIO
from typing import List, Dict, Any


def generate_proof_excel(
    calc_result: Dict[str, Any],
    extracted_json: Dict[str, Any],
    clause_type: str = "",
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation Proof"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = f"VC Waterfall — Calculation Proof ({clause_type})"
    ws["A1"].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:D3")
    ws["A3"] = "Input Parameters"
    ws["A3"].font = header_font

    row = 4
    for key, val in extracted_json.items():
        if isinstance(val, (list, dict)):
            continue
        ws.cell(row=row, column=1, value=key).border = thin_border
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.border = thin_border
        if isinstance(val, float):
            cell_val.number_format = "#,##0.00"
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Derivation Steps")
    ws.cell(row=row, column=1).font = header_font
    row += 1

    headers = ["Step", "Formula", "Values", "Result"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    derivation_steps: List[dict] = calc_result.get("derivation_steps", [])
    for step in derivation_steps:
        ws.cell(row=row, column=1, value=step.get("step", "")).border = thin_border
        ws.cell(row=row, column=2, value=step.get("formula", "")).border = thin_border
        ws.cell(row=row, column=3, value=step.get("values", "")).border = thin_border
        result_cell = ws.cell(row=row, column=4, value=step.get("result", ""))
        result_cell.border = thin_border
        if isinstance(step.get("result"), float):
            result_cell.number_format = "#,##0.0000"
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Final Results").font = header_font
    row += 1

    skip_keys = {"derivation_steps", "confidence_level", "formula_steps"}
    for key, val in calc_result.items():
        if key in skip_keys or isinstance(val, (list, dict)):
            continue
        ws.cell(row=row, column=1, value=key).border = thin_border
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.border = thin_border
        if isinstance(val, float):
            cell_val.number_format = "#,##0.0000"
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
